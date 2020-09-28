from icalendar import Calendar, Event
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
from datetime import timedelta
from pytz import UTC  # timezone
import requests
import random

# Simple program that downloads, parses and converts the data into a human-readable format on a excel spreadsheet.
# You will need an xlsx file for the program to write the calendar to: add its filepath to "fileName" variable
# Add the iCal link into the "url" variable

# ---------- EDIT TO YOUR OWN VALUES -----------
url = 'https://www.officeholidays.com/ics/united-kingdom/england'
fileName = 'CalenderiserBasicTest.xlsx'

# ---------- ----------------------- -----------

# Various functions to run the program

ls = []

now = datetime.date(UTC.localize(datetime.now()))


def run():  # Main function that runs the show
    print('--- START ---')
    parseICAL(url, 'ical.ics')
    clearCells(fileName)
    makeXL(fileName)
    addStartDates(ls, fileName)
    # dict.addEndDates(dict.ls)
    colourCells(ls, fileName)
    print('--- END ---')


def downloadICAL(name, link):  # Function to download the iCal
    try:
        r = requests.get(link, allow_redirects=True)
        open(name, 'wb').write(r.content)
        return name
    except:
        return name


def parseICAL(link, iCalfileName):  # Function that parses the iCal into a dictionary
    count = 1
    g = open(downloadICAL(iCalfileName, link), 'rb')
    gcal = Calendar.from_ical(g.read())
    for component in gcal.walk():
        if component.name == "VEVENT":
            name = str(component.get('summary')).strip()
            start = component.get('dtstart').dt
            end = component.get('dtend').dt
            ls.append({'name': name, 'start': start, 'end': end})
            count += 1
    g.close()
    return print('iCal Parsed')


def clearCells(name):  # Function to clear all the cells before the start of each run
    wb = openpyxl.load_workbook(name)
    ws = wb['Calendar']
    ws.delete_rows(1, 3)
    wb.save(name)
    return print('cells cleared')


def makeXL(name):  # Function to create the top axis of the spreadsheet with all the dates
    wb = openpyxl.load_workbook(name)
    ws = wb['Calendar']
    time = datetime.now()  # Non-pytz rated now time
    for i in range(2, 1000):
        ws.cell(row=1, column=i + 1, value=time.strftime('%d/%m/%Y'))
        time += timedelta(days=1)
    ws.cell(row=2, column=1, value='Events')
    wb.save(name)
    return print('Dates axis created')


def addStartDates(lst, name):  # Function to add the start dates to the calendar
    wb = openpyxl.load_workbook(name)
    ws = wb['Calendar']
    row = 2
    for dic in lst:
        start = dic.get('start')
        name = dic.get('name')
        for colNum in range(1, ws.max_column):
            if start.strftime('%d/%m/20%y') == ws.cell(row=1, column=colNum).value:
                ws.cell(row=row, column=colNum, value=name)
    wb.save(name)
    return print('start dates added')


def colourCells(lst, name):  # Function that colours the cells with a random colour for their duration
    wb = openpyxl.load_workbook(name)
    ws = wb['Calendar']
    for entry in lst:
        randColour = ("%06x" % random.randint(0, 0xFFFFFF))
        stayLength = entry.get('start') - entry.get('end')
        for col in range(1, ws.max_column):
            if ws.cell(row=2, column=col).value == entry.get('name'):
                for i in range(abs(stayLength.days)):
                    ws.cell(row=2, column=col + i).fill = PatternFill(fgColor=randColour, fill_type="solid")
    wb.save(name)
    return print('cells coloured')


# RUN:
run()
